"""Wingmen CTO Bot — multi-client Telegram bot for brainstorming and build jobs.

Supports:
- Musa (admin): full access to all repos, admin commands, client management
- Clients: scoped to their repos, can brainstorm and queue builds
- Free text: context-aware brainstorm using Claude with real repo knowledge
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

from dotenv import load_dotenv
from supabase import acreate_client
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from telegram.request import HTTPXRequest

import context_loader
from agents.router import build_router_prompt, parse_router_response
from agents.brainstorm import build_brainstorm_prompt
from agents.auditor import build_auditor_prompt, parse_auditor_response
from agents.fixer import build_fixer_prompt

# ── Whisper (local transcription) ────────────────────────────────
_whisper_model = None


def _get_whisper():
    global _whisper_model
    if _whisper_model is None:
        import whisper
        _whisper_model = whisper.load_model("base")
    return _whisper_model


def _whisper_transcribe(audio_path: str) -> str:
    """Transcribe audio file using local Whisper model. Runs in thread."""
    model = _get_whisper()
    result = model.transcribe(audio_path, language=None)  # auto-detect language
    return result.get("text", "").strip()


# ── Setup ────────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent / ".env")

LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "cto_bot.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("wingmen.cto_bot")

# Silence noisy loggers
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("telegram.ext").setLevel(logging.WARNING)

MUSA_TELEGRAM_ID = os.environ.get("MUSA_TELEGRAM_ID", "")

_supabase = None
_supabase_lock = asyncio.Lock()


async def get_supabase():
    global _supabase
    if _supabase is not None:
        return _supabase
    async with _supabase_lock:
        if _supabase is None:
            url = os.environ["SUPABASE_URL"]
            key = os.environ["SUPABASE_SERVICE_KEY"]
            _supabase = await acreate_client(url, key)
        return _supabase


# ── User Resolution ──────────────────────────────────────────────

# ── User Cache (with eviction) ───────────────────────────────────

_user_cache: dict[str, dict] = {}
CACHE_TTL = 300  # 5 min
CACHE_MAX_SIZE = 200


def _evict_cache():
    """Remove expired entries from user cache."""
    if len(_user_cache) <= CACHE_MAX_SIZE:
        return
    now = time.monotonic()
    expired = [k for k, v in _user_cache.items() if now - v["ts"] > CACHE_TTL]
    for k in expired:
        del _user_cache[k]
    # If still over limit, drop oldest
    if len(_user_cache) > CACHE_MAX_SIZE:
        oldest = sorted(_user_cache, key=lambda k: _user_cache[k]["ts"])
        for k in oldest[:len(_user_cache) - CACHE_MAX_SIZE]:
            del _user_cache[k]


async def resolve_user(chat_id: str) -> dict | None:
    """Resolve a Telegram chat_id to a user object. Returns None if unregistered."""
    now = time.monotonic()
    cached = _user_cache.get(chat_id)
    if cached and (now - cached["ts"]) < CACHE_TTL:
        return cached

    _evict_cache()

    if chat_id == MUSA_TELEGRAM_ID:
        user = {
            "role": "admin",
            "client": None,
            "client_id": None,
            "repos": context_loader.get_all_repo_names(),
            "name": "Musa",
            "ts": now,
        }
        _user_cache[chat_id] = user
        return user

    supabase = await get_supabase()
    result = await supabase.table("clients").select("*").eq(
        "telegram_chat_id", chat_id
    ).eq("active", True).limit(1).execute()

    if not result.data:
        return None

    client = result.data[0]
    repos = await context_loader.get_client_repos(client["id"], supabase)
    repo_names = [r["name"] for r in repos]

    # Load plan limits
    plan_id = client.get("plan", "free")
    plan_result = await supabase.table("plan_tiers").select("*").eq("id", plan_id).limit(1).execute()
    plan = plan_result.data[0] if plan_result.data else {
        "chats_per_day": 10, "builds_per_month": 2, "priority_queue": False
    }

    user = {
        "role": "client",
        "client": client,
        "client_id": client["id"],
        "repos": repo_names,
        "name": client["name"],
        "plan": plan,
        "ts": now,
    }
    _user_cache[chat_id] = user
    return user


def is_admin(user: dict) -> bool:
    return user["role"] == "admin"


# ── Onboarding ───────────────────────────────────────────────────

_onboarding_state: dict[str, dict] = {}  # chat_id -> {"step": ..., "data": {...}}


async def handle_onboarding(update: Update, chat_id: str) -> bool:
    """Smart onboarding — detects intent, auto-provisions sites.

    Flow:
    1. start → ask name
    2. name → ask what they need (or detect from initial message)
    3. intent → collect details based on type
    4. provision → create site automatically
    """
    import provisioner

    supabase = await get_supabase()
    text = (update.message.text or "").strip()

    # Check if already has a pending signup
    existing = await supabase.table("pending_signups").select("*").eq(
        "telegram_chat_id", chat_id
    ).limit(1).execute()

    if existing.data and existing.data[0]["status"] == "pending":
        await update.message.reply_text(
            "Your site is being set up! I'll message you as soon as it's ready \U0001f44d"
        )
        return True

    state = _onboarding_state.get(chat_id, {"step": "start", "data": {}})

    if state["step"] == "start":
        tg_user = update.effective_user
        _onboarding_state[chat_id] = {
            "step": "name",
            "data": {
                "telegram_username": tg_user.username or "",
                "initial_message": text,
            },
            "_ts": time.monotonic(),
        }
        await update.message.reply_text(
            "Assalamu alaikum! \U0001f44b Welcome to Wingmen.\n\n"
            "I help people create websites and manage their online presence.\n\n"
            "What's your name?"
        )
        return True

    elif state["step"] == "name":
        state["data"]["name"] = text
        state["step"] = "intent"
        _onboarding_state[chat_id] = state

        # Try to detect intent from initial message
        initial = state["data"].get("initial_message", "").lower()
        detected = _detect_intent(initial)

        if detected:
            state["data"]["template"] = detected
            state["step"] = "details"
            _onboarding_state[chat_id] = state
            return await _ask_details(update, state, detected)

        await update.message.reply_text(
            f"Nice to meet you, {text}! What can I help you with?\n\n"
            "\U0001f3a8 Create a portfolio or personal site\n"
            "\U0001f6d2 Set up an online store / ordering page\n"
            "\U0001f4c4 Build a simple landing page\n"
            "\U0001f4ac Or just tell me what you need!"
        )
        return True

    elif state["step"] == "intent":
        detected = _detect_intent(text)
        if not detected:
            await update.message.reply_text(
                "No worries! Just tell me:\n"
                "1 = Portfolio/personal site\n"
                "2 = Store/ordering page\n"
                "3 = Landing page"
            )
            return True

        state["data"]["template"] = detected
        state["step"] = "details"
        _onboarding_state[chat_id] = state
        return await _ask_details(update, state, detected)

    elif state["step"] == "details":
        template = state["data"].get("template", "portfolio")
        _parse_details(state, text, template)

        # Check if we have enough info to provision
        if _has_enough_info(state, template):
            state["step"] = "confirm"
            _onboarding_state[chat_id] = state
            return await _show_confirmation(update, state, template)
        else:
            return await _ask_missing(update, state, template)

    elif state["step"] == "confirm":
        if text.lower() in ("yes", "y", "go", "go ahead", "yep", "sure", "ok", "do it", "confirm"):
            await update.message.reply_text(
                f"\U0001f680 Setting up your site now! This takes about 2 minutes.\n"
                "I'll send you a screenshot when it's ready..."
            )
            await update.message.chat.send_action("typing")

            # Create client
            name = state["data"]["name"]
            company = state["data"].get("company", name)
            client_result = await supabase.table("clients").insert({
                "name": name,
                "telegram_chat_id": chat_id,
                "telegram_username": state["data"].get("telegram_username", ""),
                "plan": "free",
            }).execute()

            if not client_result.data:
                await update.message.reply_text("Something went wrong. Let me get the team to help.")
                return True

            client_id = client_result.data[0]["id"]

            # Clear user cache so they're recognized
            _user_cache.pop(chat_id, None)

            # Provision the site
            template = state["data"]["template"]
            config = _build_provision_config(state, template)

            result = await provisioner.provision_site(
                supabase, client_id, template, company, config,
            )

            _onboarding_state.pop(chat_id, None)

            if result["success"]:
                # Notify admin
                admin_token = os.environ.get("TELEGRAM_BOT_TOKEN")
                admin_id = os.environ.get("MUSA_TELEGRAM_ID")
                if admin_token and admin_id:
                    import httpx as hx
                    async with hx.AsyncClient(timeout=10) as client:
                        await client.post(
                            f"https://api.telegram.org/bot{admin_token}/sendMessage",
                            json={"chat_id": admin_id, "text": f"\U0001f195 New client auto-provisioned:\nName: {name}\nSite: {result['deploy_url']}\nTemplate: {template}\nRepo: {result['repo_name']}"},
                        )

                # Take screenshot and send
                deploy_url = result["deploy_url"]
                await update.message.reply_text(
                    f"\u2705 Your site is live!\n\n"
                    f"\U0001f310 {deploy_url}\n\n"
                    "You can now tell me anytime you want to:\n"
                    "\u2022 Update content or products\n"
                    "\u2022 Change colors or layout\n"
                    "\u2022 Add new pages or features\n\n"
                    "Just message me like you would a friend!"
                )

                # Try to send screenshot
                try:
                    import tempfile
                    ss_path = os.path.join(tempfile.gettempdir(), f"onboard_{client_id}.png")
                    proc = await asyncio.create_subprocess_exec(
                        "npx", "playwright", "screenshot", deploy_url, ss_path,
                        "--viewport-size", "375,812",
                        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
                    )
                    await asyncio.wait_for(proc.communicate(), timeout=30)
                    if os.path.exists(ss_path):
                        from telegram import InputFile
                        with open(ss_path, "rb") as f:
                            await update.message.reply_photo(
                                photo=InputFile(f),
                                caption=f"\U0001f4f1 Here's how your site looks!",
                            )
                        os.unlink(ss_path)
                except Exception:
                    pass
            else:
                await update.message.reply_text(
                    "I ran into an issue setting up your site. "
                    "I've flagged it to the team — they'll sort it out and message you shortly!"
                )
            return True

        elif text.lower() in ("no", "n", "cancel", "nah"):
            _onboarding_state.pop(chat_id, None)
            await update.message.reply_text("No problem! Message me anytime you're ready.")
            return True
        else:
            await update.message.reply_text("Just say 'yes' to confirm or 'no' to cancel.")
            return True

    return False


def _detect_intent(text: str) -> str | None:
    """Detect what type of site the user wants from their message."""
    text = text.lower()
    storefront_words = ["store", "shop", "sell", "order", "menu", "food", "bake", "kueh", "business", "product", "inventory", "price"]
    portfolio_words = ["portfolio", "personal", "designer", "photographer", "freelance", "showcase", "work", "cv", "resume"]
    landing_words = ["landing", "launch", "coming soon", "waitlist", "startup"]

    store_score = sum(1 for w in storefront_words if w in text)
    portfolio_score = sum(1 for w in portfolio_words if w in text)
    landing_score = sum(1 for w in landing_words if w in text)

    if text in ("1", "portfolio"):
        return "portfolio"
    if text in ("2", "store", "storefront"):
        return "storefront"
    if text in ("3", "landing"):
        return "landing"

    if store_score > portfolio_score and store_score > landing_score:
        return "storefront"
    if portfolio_score > store_score and portfolio_score > landing_score:
        return "portfolio"
    if landing_score > 0:
        return "landing"
    return None


async def _ask_details(update, state, template):
    name = state["data"]["name"]
    if template == "portfolio":
        await update.message.reply_text(
            f"A portfolio site — great choice, {name}!\n\n"
            "Tell me about yourself in one message:\n"
            "\u2022 What do you do? (design, photography, dev, etc.)\n"
            "\u2022 Your brand/business name\n"
            "\u2022 Color preference? (or I'll pick something clean)\n"
            "\u2022 Your email for the contact section"
        )
    elif template == "storefront":
        await update.message.reply_text(
            f"An online store — love it, {name}!\n\n"
            "Tell me about your business in one message:\n"
            "\u2022 Business name\n"
            "\u2022 What do you sell? (list items with prices)\n"
            "\u2022 Your WhatsApp/phone number\n"
            "\u2022 PayNow UEN or phone (for payments)"
        )
    else:
        await update.message.reply_text(
            f"A landing page — nice, {name}!\n\n"
            "Tell me:\n"
            "\u2022 Project/company name\n"
            "\u2022 One-line description\n"
            "\u2022 Color preference?"
        )
    return True


def _parse_details(state, text, template):
    """Parse user's details message and extract what we can."""
    state["data"]["details_raw"] = text
    state["data"]["company"] = state["data"].get("company", state["data"]["name"])

    # Try to extract email
    import re
    email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', text)
    if email_match:
        state["data"]["email"] = email_match.group()

    # Try to extract phone
    phone_match = re.search(r'[\+]?[0-9\s\-]{8,15}', text)
    if phone_match:
        state["data"]["phone"] = phone_match.group().strip()

    # Try to extract prices/products for storefront
    if template == "storefront":
        products = []
        # Pattern: "item name $price" or "item name price"
        price_matches = re.findall(r'([a-zA-Z\s]+?)[\s]*[\$S]*(\d+(?:\.\d{2})?)', text)
        for name, price in price_matches:
            name = name.strip().strip(',').strip('-').strip()
            if name and len(name) > 1:
                products.append({"name": name, "price": float(price)})
        if products:
            state["data"]["products"] = products

    # Use the raw text as company name if we don't have one
    lines = text.strip().splitlines()
    if lines:
        first_line = lines[0].strip()
        if len(first_line) < 50 and not first_line.startswith(("I ", "My ", "We ")):
            state["data"]["company"] = first_line


def _has_enough_info(state, template):
    """Check if we have enough to provision."""
    data = state["data"]
    if template == "storefront":
        return bool(data.get("company")) and bool(data.get("details_raw"))
    return bool(data.get("company")) and bool(data.get("details_raw"))


async def _ask_missing(update, state, template):
    """This shouldn't normally be called — we accept whatever they give."""
    await update.message.reply_text("Got it! Anything else to add? Or say 'done' and I'll set it up.")
    state["step"] = "confirm"
    return True


async def _show_confirmation(update, state, template):
    data = state["data"]
    name = data["name"]
    company = data.get("company", name)

    if template == "storefront" and data.get("products"):
        products_list = "\n".join(f"  \u2022 {p['name']} — ${p['price']:.2f}" for p in data["products"])
        await update.message.reply_text(
            f"Here's what I'll set up for you:\n\n"
            f"\U0001f6d2 Online Store: {company}\n"
            f"Products:\n{products_list}\n"
            f"Phone: {data.get('phone', 'will add later')}\n\n"
            "Your store will be live in about 2 minutes.\n\n"
            "Should I go ahead?"
        )
    elif template == "portfolio":
        await update.message.reply_text(
            f"Here's what I'll set up:\n\n"
            f"\U0001f3a8 Portfolio: {company}\n"
            f"Contact: {data.get('email', 'will add later')}\n\n"
            "A clean, mobile-friendly portfolio site ready in 2 minutes.\n\n"
            "Should I go ahead?"
        )
    else:
        await update.message.reply_text(
            f"Here's what I'll set up:\n\n"
            f"\U0001f4c4 Landing Page: {company}\n\n"
            "Ready in about 2 minutes. Should I go ahead?"
        )
    return True


def _build_provision_config(state, template):
    """Build the config dict for provisioner from onboarding state."""
    data = state["data"]
    config = {
        "tagline": "",
        "description": data.get("details_raw", "")[:200],
        "primary_color": "#111111" if template == "portfolio" else "#2d6a4f",
    }

    if template == "portfolio":
        config["email"] = data.get("email", "")
    elif template == "storefront":
        config["phone"] = data.get("phone", "")
        config["paynow_uen"] = data.get("paynow_uen", "")
        config["paynow_name"] = data.get("company", "").upper()
        config["products"] = data.get("products", [])

    return config


async def check_usage_limit(user: dict, action: str) -> str | None:
    """Check if user is within their plan limits. Returns error message or None if OK."""
    if is_admin(user):
        return None

    plan = user.get("plan", {})
    client_id = user.get("client_id")
    if not client_id:
        return None

    supabase = await get_supabase()

    if action == "chat":
        # Check daily chat limit
        from datetime import datetime, timezone, timedelta
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0).isoformat()
        result = await (
            supabase.table("usage_log")
            .select("id", count="exact")
            .eq("client_id", client_id)
            .eq("action_type", "chat")
            .gte("created_at", today_start)
            .execute()
        )
        count = result.count if hasattr(result, 'count') and result.count else len(result.data or [])
        limit = plan.get("chats_per_day", 10)
        if count >= limit:
            return (
                f"You've reached your daily chat limit ({limit} messages). "
                "Upgrade your plan for more! Send /upgrade to see options."
            )

    elif action == "build":
        # Check monthly build limit
        client_data = user.get("client", {})
        monthly_builds = client_data.get("monthly_build_count", 0)
        limit = plan.get("builds_per_month", 2)
        if monthly_builds >= limit:
            return (
                f"You've used all {limit} builds this month. "
                "Upgrade your plan for more! Send /upgrade to see options."
            )

    return None


# ── Per-user state ───────────────────────────────────────────────

_active_repo: dict[str, str] = {}            # chat_id -> repo_name
_rate_limits: dict[str, list[float]] = {}    # chat_id -> list of timestamps
_pending_fixes: dict[str, dict] = {}         # chat_id -> {"issues": [...], "repo": str}
_admin_session_id: str | None = None         # persistent Claude session for admin
_admin_session_ts: float = 0                 # last activity timestamp
ADMIN_SESSION_TTL = 3600                     # 1 hour inactivity before new session

MAX_BUILDS_PER_HOUR = 10
MAX_MSG_LENGTH = 2000
_message_counter = 0


def _cleanup_caches():
    """Periodic cleanup of unbounded caches. Called every 100 messages."""
    global _message_counter
    _message_counter += 1
    if _message_counter < 100:
        return
    _message_counter = 0

    now = time.monotonic()

    # _onboarding_state: remove entries older than 1 hour
    stale_onboarding = [
        k for k, v in _onboarding_state.items()
        if now - v.get("_ts", now) > 3600
    ]
    for k in stale_onboarding:
        del _onboarding_state[k]

    # _rate_limits: remove empty lists
    empty_rate = [k for k, v in _rate_limits.items() if not v]
    for k in empty_rate:
        del _rate_limits[k]

    # _active_repo: cap at 200 entries (drop oldest by insertion order)
    if len(_active_repo) > 200:
        keys = list(_active_repo.keys())
        for k in keys[:len(_active_repo) - 200]:
            del _active_repo[k]


def check_rate_limit(chat_id: str) -> bool:
    """Returns True if within rate limit, False if exceeded."""
    now = time.monotonic()
    if chat_id not in _rate_limits:
        _rate_limits[chat_id] = []
    _rate_limits[chat_id] = [t for t in _rate_limits[chat_id] if now - t < 3600]
    if len(_rate_limits[chat_id]) >= MAX_BUILDS_PER_HOUR:
        return False
    _rate_limits[chat_id].append(now)
    return True


async def get_history(chat_id: str) -> list[dict]:
    """Load chat history from Supabase, falling back to empty."""
    try:
        supabase = await get_supabase()
        result = await (
            supabase.table("chat_history")
            .select("role, content")
            .eq("chat_id", chat_id)
            .order("created_at", desc=True)
            .limit(20)
            .execute()
        )
        if result.data:
            # Reverse to chronological order
            return list(reversed(result.data))
    except Exception as e:
        logger.warning(f"Failed to load chat history: {e}")
    return []


async def save_message(chat_id: str, role: str, content: str) -> None:
    """Persist a chat message to Supabase."""
    try:
        supabase = await get_supabase()
        await supabase.table("chat_history").insert({
            "chat_id": chat_id,
            "role": role,
            "content": content[:4000],
        }).execute()
    except Exception as e:
        logger.warning(f"Failed to save chat message: {e}")


async def log_audit(chat_id: str, user_name: str, action: str, target: str = "", detail: str = "") -> None:
    """Write to audit log for all significant actions."""
    try:
        supabase = await get_supabase()
        await supabase.table("audit_log").insert({
            "user_chat_id": chat_id,
            "user_name": user_name,
            "action": action,
            "target": target,
            "detail": detail[:1000],
        }).execute()
    except Exception as e:
        logger.warning(f"Failed to write audit log: {e}")


async def log_usage(client_id: int | None, action_type: str, repo_name: str = "", tokens: int = 0, duration: float = 0) -> None:
    """Track usage for metering/billing."""
    try:
        supabase = await get_supabase()
        await supabase.table("usage_log").insert({
            "client_id": client_id,
            "action_type": action_type,
            "repo_name": repo_name,
            "tokens_used": tokens,
            "duration_seconds": duration,
        }).execute()
    except Exception as e:
        logger.warning(f"Failed to log usage: {e}")


async def _load_active_repo(chat_id: str) -> str | None:
    """Try to infer active repo from recent chat history."""
    try:
        supabase = await get_supabase()
        result = await (
            supabase.table("chat_history")
            .select("content")
            .eq("chat_id", chat_id)
            .eq("role", "user")
            .order("created_at", desc=True)
            .limit(5)
            .execute()
        )
        if result.data:
            all_repos = context_loader.get_all_repo_names()
            for msg in result.data:
                for repo in all_repos:
                    if repo.lower() in msg["content"].lower():
                        return repo
    except Exception:
        pass
    return None


async def _queue_message(chat_id: str, message_id: int, msg_type: str, content: str = "", file_id: str = "") -> None:
    """Save incoming message to queue before processing. Idempotent via unique constraint."""
    try:
        supabase = await get_supabase()
        await supabase.table("message_queue").upsert({
            "chat_id": chat_id,
            "message_id": message_id,
            "message_type": msg_type,
            "content": content[:4000] if content else "",
            "file_id": file_id,
            "processed": False,
        }, on_conflict="chat_id,message_id").execute()
    except Exception as e:
        logger.warning(f"Failed to queue message: {e}")


async def _mark_processed(chat_id: str, message_id: int) -> None:
    """Mark a queued message as processed."""
    try:
        supabase = await get_supabase()
        await supabase.table("message_queue").update({"processed": True}).eq(
            "chat_id", chat_id
        ).eq("message_id", message_id).execute()
    except Exception as e:
        logger.warning(f"Failed to mark message processed: {e}")


async def _recover_unprocessed(app) -> None:
    """On startup, find unprocessed messages and notify those users."""
    try:
        supabase = await get_supabase()
        # Only recover messages from the last hour (older ones are stale)
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
        result = await (
            supabase.table("message_queue")
            .select("chat_id, content, message_type")
            .eq("processed", False)
            .gte("created_at", cutoff)
            .order("created_at", desc=False)
            .limit(50)
            .execute()
        )

        if not result.data:
            return

        # Group by chat_id and send apology + last message summary
        from collections import defaultdict
        by_chat: dict[str, list] = defaultdict(list)
        for msg in result.data:
            by_chat[msg["chat_id"]].append(msg)

        for chat_id, msgs in by_chat.items():
            count = len(msgs)
            last_content = msgs[-1].get("content", "")[:100]
            try:
                await app.bot.send_message(
                    chat_id=int(chat_id),
                    text=(
                        f"Sorry, I missed {count} message{'s' if count > 1 else ''} "
                        f"during a brief restart. Could you resend?\n\n"
                        f"Last message: \"{last_content}...\""
                    ),
                )
            except Exception:
                pass

        # Mark all as processed
        for msg in result.data:
            await supabase.table("message_queue").update({"processed": True}).eq(
                "chat_id", msg["chat_id"]
            ).execute()

        logger.info(f"Recovered {len(result.data)} unprocessed messages across {len(by_chat)} chats")

        # Clean up processed messages older than 24 hours
        cleanup_cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        try:
            await (
                supabase.table("message_queue")
                .delete()
                .eq("processed", True)
                .lt("created_at", cleanup_cutoff)
                .execute()
            )
            logger.info("Cleaned up processed messages older than 24 hours")
        except Exception as cleanup_err:
            logger.warning(f"Message queue cleanup failed: {cleanup_err}")
    except Exception as e:
        logger.error(f"Message recovery failed: {e}")


def get_active_repo(chat_id: str, user: dict) -> str | None:
    if chat_id in _active_repo:
        return _active_repo[chat_id]
    if len(user["repos"]) == 1:
        _active_repo[chat_id] = user["repos"][0]
        return user["repos"][0]
    return None


# ── Commands ─────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)

    if not user:
        await handle_onboarding(update, chat_id)
        return

    if is_admin(user):
        await update.message.reply_text(
            "Assalamu alaikum, CTO.\n\n"
            "Commands:\n"
            "/build <repo> <task> — Queue a build job\n"
            "/status [repo] — Show status\n"
            "/jobs — Show queued/running jobs\n"
            "/repos — List all repos\n"
            "/repo <name> — Switch active repo\n"
            "/pause <job_id> — Pause a job\n"
            "/cancel <job_id> — Cancel a job\n"
            "/priority <job_id> <1-5> — Set job priority\n"
            "/addclient <name> <chat_id> <plan> — Register client\n"
            "/linkrepo <client_name> <repo_name> — Link repo to client\n"
            "/clients — List all clients\n"
            "/id — Show your Telegram chat ID\n\n"
            "Or just type normally to brainstorm."
        )
    else:
        repo = get_active_repo(chat_id, user)
        if len(user["repos"]) == 1:
            _active_repo[chat_id] = user["repos"][0]
            repo = user["repos"][0]

        if repo:
            await update.message.reply_text(
                f"Assalamu alaikum, {user['name']}! \U0001f44b\n\n"
                f"I'm your project assistant for {repo}.\n\n"
                "Just tell me what you need — whether it's:\n"
                "\u2022 Updating content or prices\n"
                "\u2022 Fixing something that's not working\n"
                "\u2022 Adding a new feature\n"
                "\u2022 Checking on progress\n\n"
                "No technical knowledge needed — just describe what you want in your own words!"
            )
        else:
            repo_list = "\n".join(f"\u2022 {r}" for r in user["repos"])
            await update.message.reply_text(
                f"Assalamu alaikum, {user['name']}! \U0001f44b\n\n"
                "I'm your project assistant from Wingmen.\n\n"
                f"You have these projects:\n{repo_list}\n\n"
                "Which one would you like to work on? Just mention its name or type /repo to switch."
            )


async def cmd_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Your chat ID: {update.effective_user.id}")


async def cmd_upgrade(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show plan options."""
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)

    current_plan = "free"
    if user and user.get("client"):
        current_plan = user["client"].get("plan", "free")

    plans = (
        "Your plan: " + current_plan.upper() + "\n\n"
        "\U0001f7e2 FREE — $0/mo\n"
        "  10 chats/day, 2 builds/month\n\n"
        "\U0001f535 STARTER — $49/mo\n"
        "  50 chats/day, 10 builds/month\n\n"
        "\U0001f7e1 GROWTH — $149/mo\n"
        "  Unlimited chats, 30 builds/month, priority\n\n"
        "\U0001f7e0 SCALE — $399/mo\n"
        "  Unlimited everything, dedicated support\n\n"
        "To upgrade, contact the team or send /pay <plan>"
    )
    await update.message.reply_text(plans)


async def cmd_plan_usage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show realtime plan usage for admin — percentage of Max subscription used."""
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return

    supabase = await get_supabase()
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    # Today's CLI calls (builds + chats + specs)
    day_result = await supabase.table("usage_log").select("id, action_type, duration_seconds").gte(
        "created_at", today_start
    ).limit(1000).execute()

    day_chats = sum(1 for r in (day_result.data or []) if r["action_type"] == "chat")
    day_builds = sum(1 for r in (day_result.data or []) if r["action_type"] in ("build_queued", "build_completed"))
    day_duration = sum(r.get("duration_seconds", 0) for r in (day_result.data or []))

    # This month's builds
    month_result = await supabase.table("usage_log").select("id, action_type, duration_seconds").gte(
        "created_at", month_start
    ).limit(1000).execute()

    month_chats = sum(1 for r in (month_result.data or []) if r["action_type"] == "chat")
    month_builds = sum(1 for r in (month_result.data or []) if r["action_type"] in ("build_queued", "build_completed"))
    month_duration = sum(r.get("duration_seconds", 0) for r in (month_result.data or []))

    # Active clients
    clients_result = await supabase.table("clients").select("id").eq("active", True).execute()
    active_clients = len(clients_result.data or [])

    # Active jobs
    jobs_result = await supabase.table("jobs").select("id, status").in_(
        "status", ["queued", "running"]
    ).execute()
    queued = sum(1 for j in (jobs_result.data or []) if j["status"] == "queued")
    running = sum(1 for j in (jobs_result.data or []) if j["status"] == "running")

    # Estimate plan usage (Pro 5x ~= 500 messages/day equivalent)
    # Each CLI call ≈ 1 message. Builds use ~10-50 messages each.
    estimated_daily_messages = day_chats + (day_builds * 30)
    daily_limit = 500  # rough Pro 5x estimate
    pct = min(100, int((estimated_daily_messages / daily_limit) * 100))

    bar_filled = pct // 5
    bar_empty = 20 - bar_filled
    bar = "\u2588" * bar_filled + "\u2591" * bar_empty

    days_left = (now.replace(month=now.month % 12 + 1, day=1) - now).days

    msg = (
        f"\U0001f4ca Plan Usage\n\n"
        f"Daily: [{bar}] {pct}%\n"
        f"  {day_chats} chats, {day_builds} builds, {day_duration:.0f}s CLI time\n\n"
        f"Monthly ({days_left} days left):\n"
        f"  {month_chats} chats, {month_builds} builds\n"
        f"  {month_duration / 3600:.1f}h total CLI time\n\n"
        f"System:\n"
        f"  {active_clients} active clients\n"
        f"  {running} running / {queued} queued jobs"
    )
    await update.message.reply_text(msg)


async def cmd_undo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Undo the last build — git revert + redeploy."""
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    repo_name = get_active_repo(chat_id, user)
    if not repo_name:
        await update.message.reply_text("Select a repo first with /repo <name>")
        return

    try:
        config = context_loader.get_repo_config(repo_name)
        repo_path = os.path.expanduser(config["local_path"])

        # Check last commit
        proc = await asyncio.create_subprocess_exec(
            "git", "log", "--oneline", "-1",
            cwd=repo_path,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await proc.communicate()
        last_commit = stdout.decode(errors="replace").strip()

        if not last_commit:
            await update.message.reply_text("No commits to undo.")
            return

        # Confirm with user
        if not context.args or context.args[0] != "--confirm":
            await update.message.reply_text(
                f"Last commit on {repo_name}:\n{last_commit}\n\n"
                "This will revert it and redeploy. Send /undo --confirm to proceed."
            )
            return

        # Revert
        proc = await asyncio.create_subprocess_exec(
            "git", "revert", "HEAD", "--no-edit",
            cwd=repo_path,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        _, stderr = await proc.communicate()
        if proc.returncode != 0:
            await update.message.reply_text(f"Revert failed: {stderr.decode(errors='replace')[:500]}")
            return

        # Push
        proc = await asyncio.create_subprocess_exec(
            "git", "push",
            cwd=repo_path,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        await asyncio.wait_for(proc.communicate(), timeout=60)

        await log_audit(chat_id, user["name"], "undo", repo_name, last_commit)
        await update.message.reply_text(f"\u21a9\ufe0f Reverted and pushed: {last_commit}\n\nVercel will redeploy automatically.")

    except Exception as e:
        logger.error(f"Undo failed: {e}")
        await update.message.reply_text("Undo failed. Check with the team.")


async def cmd_preview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Take a screenshot of the live site and send it."""
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    repo_name = get_active_repo(chat_id, user)
    if not repo_name:
        await update.message.reply_text("Select a repo first with /repo <name>")
        return

    try:
        config = context_loader.get_repo_config(repo_name)
        deploy_url = config.get("deploy_url")
        if not deploy_url or deploy_url == "FILL_IN":
            await update.message.reply_text(f"No deploy URL configured for {repo_name}.")
            return

        await update.message.reply_text(f"\U0001f4f8 Taking screenshot of {deploy_url}...")

        # Use a headless browser screenshot via CLI
        import tempfile
        screenshot_path = os.path.join(tempfile.gettempdir(), f"preview_{repo_name}.png")

        # Try playwright first, fall back to message
        proc = await asyncio.create_subprocess_exec(
            "npx", "playwright", "screenshot", deploy_url, screenshot_path,
            "--viewport-size", "375,812",  # iPhone viewport
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)

        if proc.returncode == 0 and os.path.exists(screenshot_path):
            from telegram import InputFile
            with open(screenshot_path, "rb") as f:
                await update.message.reply_photo(photo=InputFile(f), caption=f"\U0001f4f1 {repo_name} — {deploy_url}")
            os.unlink(screenshot_path)
        else:
            await update.message.reply_text(f"Your site is live at: {deploy_url}")

    except asyncio.TimeoutError:
        await update.message.reply_text(f"Screenshot timed out. Your site is at: {deploy_url}")
    except Exception as e:
        logger.error(f"Preview failed: {e}")
        config = context_loader.get_repo_config(repo_name)
        await update.message.reply_text(f"Your site is live at: {config.get('deploy_url', 'N/A')}")


async def cmd_digest(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show weekly digest of changes for active repo."""
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    repo_name = get_active_repo(chat_id, user)
    if not repo_name:
        await update.message.reply_text("Select a repo first with /repo <name>")
        return

    try:
        config = context_loader.get_repo_config(repo_name)
        repo_path = os.path.expanduser(config["local_path"])

        # Get last 7 days of commits
        proc = await asyncio.create_subprocess_exec(
            "git", "log", "--oneline", "--since=7 days ago",
            cwd=repo_path,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await proc.communicate()
        commits = stdout.decode(errors="replace").strip()

        # Get completed jobs this week
        supabase = await get_supabase()
        week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
        jobs_result = await (
            supabase.table("jobs")
            .select("id, description, status, created_at")
            .eq("repo_name", repo_name)
            .eq("status", "completed")
            .gte("created_at", week_ago)
            .order("created_at", desc=True)
            .limit(10)
            .execute()
        )

        lines = [f"\U0001f4cb Weekly Digest — {repo_name}\n"]

        if jobs_result.data:
            lines.append("Completed tasks:")
            for j in jobs_result.data:
                desc = j["description"][:100].split("\n")[0]
                lines.append(f"  \u2705 #{j['id']}: {desc}")
        else:
            lines.append("No completed tasks this week.")

        if commits:
            lines.append(f"\nCommits ({commits.count(chr(10)) + 1}):")
            for line in commits.splitlines()[:10]:
                lines.append(f"  {line}")
        else:
            lines.append("\nNo commits this week.")

        await update.message.reply_text("\n".join(lines))

    except Exception as e:
        logger.error(f"Digest failed: {e}")
        await update.message.reply_text("Couldn't generate digest. Try again later.")


async def cmd_screenshots(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Take screenshots of specific routes and send them.

    Usage: /screenshots [route1 route2 ...]
    Example: /screenshots /admin/qurban /donor/qurban
    No args = screenshot all known routes for the active repo.
    """
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    repo_name = get_active_repo(chat_id, user)
    if not repo_name:
        await update.message.reply_text("Select a repo first with /repo <name>")
        return

    try:
        config = context_loader.get_repo_config(repo_name)
        deploy_url = config.get("deploy_url", "")
        if not deploy_url or deploy_url == "FILL_IN":
            await update.message.reply_text(f"No deploy URL for {repo_name}.")
            return

        # Determine which routes to screenshot
        routes = list(context.args) if context.args else []

        if not routes:
            # Auto-discover routes from the repo's file structure
            repo_path = os.path.expanduser(config["local_path"])
            app_dir = os.path.join(repo_path, "app")
            if os.path.exists(app_dir):
                for root, dirs, files in os.walk(app_dir):
                    # Skip internal Next.js dirs
                    dirs[:] = [d for d in dirs if not d.startswith("_") and not d.startswith(".")]
                    if "page.tsx" in files or "page.jsx" in files or "page.ts" in files:
                        route = root.replace(app_dir, "").replace("\\", "/")
                        # Skip dynamic routes with brackets for auto-discovery
                        if "[" not in route and route:
                            routes.append(route)
            if not routes:
                routes = ["/"]

        await update.message.reply_text(f"\U0001f4f8 Taking {len(routes)} screenshot{'s' if len(routes) > 1 else ''}...")

        import tempfile
        base_url = deploy_url.rstrip("/")

        for route in routes:
            route = route if route.startswith("/") else f"/{route}"
            url = f"{base_url}{route}"
            screenshot_path = os.path.join(tempfile.gettempdir(), f"ss_{repo_name}_{route.replace('/', '_')}.png")

            # Desktop for admin routes, mobile for others
            # Default to mobile — most users view on phone
            viewport = "375,812"

            try:
                proc = await asyncio.create_subprocess_exec(
                    "npx", "playwright", "screenshot", url, screenshot_path,
                    "--viewport-size", viewport,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                )
                await asyncio.wait_for(proc.communicate(), timeout=30)

                if proc.returncode == 0 and os.path.exists(screenshot_path):
                    caption = f"\U0001f4f1 {repo_name}{route}\n{url}"
                    from telegram import InputFile
                    with open(screenshot_path, "rb") as f:
                        await update.message.reply_photo(
                            photo=InputFile(f),
                            caption=caption,
                        )
                else:
                    await update.message.reply_text(f"Failed to screenshot {route}")

            except asyncio.TimeoutError:
                await update.message.reply_text(f"Timeout on {route}")
            finally:
                try:
                    os.unlink(screenshot_path)
                except Exception:
                    pass

    except Exception as e:
        logger.error(f"Screenshots failed: {e}")
        await update.message.reply_text("Screenshot capture failed. Try again.")


async def cmd_repo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        await update.message.reply_text("Not registered.")
        return

    if not context.args:
        current = get_active_repo(chat_id, user)
        lines = [f"Active: {current or 'none'}\n", "Your repos:"]
        for r in user["repos"]:
            marker = " (active)" if r == current else ""
            lines.append(f"  - {r}{marker}")
        await update.message.reply_text("\n".join(lines))
        return

    name = context.args[0]
    if name not in user["repos"]:
        await update.message.reply_text(
            f"Unknown repo: {name}\nYour repos: {', '.join(user['repos'])}"
        )
        return

    _active_repo[chat_id] = name
    await update.message.reply_text(f"Switched to {name}")


async def cmd_build(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        await update.message.reply_text("Not registered.")
        return

    args = context.args or []

    if is_admin(user):
        # Admin: /build <repo> <task>
        if len(args) < 2:
            await update.message.reply_text("Usage: /build <repo> <task description>")
            return
        repo_name = args[0]
        description = " ".join(args[1:])
        if repo_name not in user["repos"]:
            await update.message.reply_text(
                f"Unknown repo: {repo_name}\nAvailable: {', '.join(user['repos'])}"
            )
            return
    else:
        # Client: /build <task> (uses active repo)
        if not args:
            await update.message.reply_text("Usage: /build <describe what you need>")
            return
        repo_name = get_active_repo(chat_id, user)
        if not repo_name:
            await update.message.reply_text("Select a repo first with /repo <name>")
            return
        description = " ".join(args)

    # Validate
    if len(description) > MAX_MSG_LENGTH:
        await update.message.reply_text(f"Description too long (max {MAX_MSG_LENGTH} chars).")
        return

    if not check_rate_limit(chat_id):
        await update.message.reply_text("Too many requests. Please wait a bit before submitting more.")
        return

    try:
        config = context_loader.get_repo_config(repo_name)
    except ValueError:
        await update.message.reply_text(f"Repo {repo_name} not configured.")
        return

    supabase = await get_supabase()
    result = await supabase.table("jobs").insert({
        "repo_name": repo_name,
        "description": description,
        "status": "queued",
        "priority": config["priority"],
        "triggered_by": "telegram",
        "client_id": user.get("client_id"),
    }).execute()

    if not result.data:
        await update.message.reply_text("Failed to queue job. Please try again.")
        return

    job = result.data[0]
    await update.message.reply_text(
        f"\u2705 Job #{job['id']} queued\n"
        f"\U0001f4e6 Repo: {repo_name}\n"
        f"\U0001f4dd Task: {description}\n"
        f"\u23f3 Priority: {config['priority']}"
    )
    await log_audit(chat_id, user["name"], "queue_build", repo_name, description[:200])
    await log_usage(user.get("client_id"), "build_queued", repo_name)
    logger.info(f"Job #{job['id']} queued by {user['name']}: {repo_name} — {description}")


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    # Determine which repo to show
    if context.args and is_admin(user):
        repo_name = context.args[0]
    else:
        repo_name = get_active_repo(chat_id, user)

    if repo_name:
        try:
            config = context_loader.get_repo_config(repo_name)
            repo_path = Path(os.path.expanduser(config["local_path"]))
            status_file = repo_path / "STATUS.md"
            if status_file.exists():
                content = status_file.read_text()
                if len(content) > 4000:
                    content = content[:4000] + "\n...(truncated)"
                await update.message.reply_text(content)
                return
        except ValueError:
            pass

    # Fallback: orchestrator status
    orch_status = Path(__file__).parent / "STATUS.md"
    if orch_status.exists():
        content = orch_status.read_text()
        if len(content) > 4000:
            content = content[:4000] + "\n...(truncated)"
        await update.message.reply_text(content)
    else:
        await update.message.reply_text("No status available.")


async def cmd_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    supabase = await get_supabase()
    query = (
        supabase.table("jobs")
        .select("id, repo_name, description, status, priority, fail_count, created_at")
        .in_("status", ["queued", "running", "paused"])
        .order("priority", desc=False)
    )

    # Clients only see their own jobs
    if not is_admin(user) and user["client_id"]:
        query = query.eq("client_id", user["client_id"])

    result = await query.execute()

    if not result.data:
        await update.message.reply_text("No active jobs.")
        return

    lines = ["Active Jobs:\n"]
    for j in result.data:
        icon = {"queued": "\u23f3", "running": "\u25b6\ufe0f", "paused": "\u23f8"}.get(j["status"], "\u2753")
        lines.append(f"{icon} #{j['id']} [{j['status']}] P{j['priority']} {j['repo_name']}\n   {j['description']}")
        if j["fail_count"] > 0:
            lines.append(f"   \u26a0\ufe0f Failures: {j['fail_count']}")

    await update.message.reply_text("\n".join(lines))


async def cmd_repos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        return

    if is_admin(user):
        repos = context_loader._load_repos()
    else:
        supabase = await get_supabase()
        repos = await context_loader.get_client_repos(user["client_id"], supabase)

    if not repos:
        await update.message.reply_text("No repos linked to your account.")
        return

    lines = ["Your repos:\n"]
    for r in repos:
        icon = "\U0001f7e2" if r["status"] == "active" else "\U0001f7e1"
        lines.append(f"{icon} P{r['priority']} {r['name']} — {r['status']}")
        if r.get("deploy_url") and r["deploy_url"] != "FILL_IN":
            lines.append(f"   {r['deploy_url']}")

    await update.message.reply_text("\n".join(lines))


# ── Admin-only commands ──────────────────────────────────────────

async def cmd_pause(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return
    if not context.args:
        await update.message.reply_text("Usage: /pause <job_id>")
        return
    job_id = int(context.args[0])
    supabase = await get_supabase()
    await supabase.table("jobs").update({"status": "paused"}).eq("id", job_id).execute()
    await log_audit(str(update.effective_user.id), user["name"], "pause_job", f"job_{job_id}")
    await update.message.reply_text(f"\u23f8 Job #{job_id} paused.")


async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return
    if not context.args:
        await update.message.reply_text("Usage: /cancel <job_id>")
        return
    job_id = int(context.args[0])
    supabase = await get_supabase()
    await supabase.table("jobs").update({"status": "failed", "result_summary": "Cancelled by admin"}).eq("id", job_id).execute()
    await log_audit(str(update.effective_user.id), user["name"], "cancel_job", f"job_{job_id}")
    await update.message.reply_text(f"\u274c Job #{job_id} cancelled.")


async def cmd_priority(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return
    if not context.args or len(context.args) < 2:
        await update.message.reply_text("Usage: /priority <job_id> <1-5>")
        return
    job_id = int(context.args[0])
    priority = max(1, min(5, int(context.args[1])))
    supabase = await get_supabase()
    await supabase.table("jobs").update({"priority": priority}).eq("id", job_id).execute()
    await update.message.reply_text(f"\U0001f4ca Job #{job_id} priority set to {priority}.")


async def cmd_addclient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return
    if not context.args or len(context.args) < 3:
        await update.message.reply_text("Usage: /addclient <name> <telegram_chat_id> <plan>")
        return

    name = context.args[0]
    chat_id = context.args[1]
    plan = context.args[2]

    supabase = await get_supabase()
    result = await supabase.table("clients").insert({
        "name": name,
        "telegram_chat_id": chat_id,
        "plan": plan,
    }).execute()

    if not result.data:
        await update.message.reply_text("Failed to add client. Please try again.")
        return

    client = result.data[0]
    # Clear cache so new client is recognized
    _user_cache.pop(chat_id, None)
    await log_audit(str(update.effective_user.id), user["name"], "add_client", name, f"plan={plan}, chat_id={chat_id}")
    await update.message.reply_text(f"\u2705 Client '{name}' added (id: {client['id']}, plan: {plan})")
    logger.info(f"Client added: {name} (chat_id: {chat_id}, plan: {plan})")


async def cmd_linkrepo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return
    if not context.args or len(context.args) < 2:
        await update.message.reply_text("Usage: /linkrepo <client_name> <repo_name>")
        return

    client_name = context.args[0]
    repo_name = context.args[1]

    # Validate repo exists
    try:
        context_loader.get_repo_config(repo_name)
    except ValueError:
        await update.message.reply_text(f"Repo '{repo_name}' not in REPOS.json")
        return

    supabase = await get_supabase()
    # Find client by name
    result = await supabase.table("clients").select("id, telegram_chat_id").eq("name", client_name).limit(1).execute()
    if not result.data:
        await update.message.reply_text(f"Client '{client_name}' not found")
        return

    client = result.data[0]
    await supabase.table("client_repos").insert({
        "client_id": client["id"],
        "repo_name": repo_name,
    }).execute()

    # Clear cache
    _user_cache.pop(client.get("telegram_chat_id", ""), None)
    await log_audit(str(update.effective_user.id), user["name"], "link_repo", f"{client_name}/{repo_name}")
    await update.message.reply_text(f"\u2705 Linked {repo_name} to {client_name}")
    logger.info(f"Repo {repo_name} linked to client {client_name}")


async def cmd_clients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return

    supabase = await get_supabase()
    result = await supabase.table("clients").select("id, name, telegram_chat_id, plan, active").execute()

    if not result.data:
        await update.message.reply_text("No clients registered.")
        return

    # Batch fetch all client_repos in one query instead of N+1
    client_ids = [c["id"] for c in result.data]
    all_repos_result = await supabase.table("client_repos").select("client_id, repo_name").in_("client_id", client_ids).execute()
    from collections import defaultdict
    repos_by_client: dict[int, list[str]] = defaultdict(list)
    for r in (all_repos_result.data or []):
        repos_by_client[r["client_id"]].append(r["repo_name"])

    lines = ["Clients:\n"]
    for c in result.data:
        status = "\U0001f7e2" if c["active"] else "\u26d4"
        repo_names = repos_by_client.get(c["id"], [])
        repos_str = ", ".join(repo_names) if repo_names else "none"
        lines.append(f"{status} {c['name']} [{c['plan']}] — repos: {repos_str}")

    await update.message.reply_text("\n".join(lines))


async def cmd_usage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show usage stats per client (admin only)."""
    user = await resolve_user(str(update.effective_user.id))
    if not user or not is_admin(user):
        return

    supabase = await get_supabase()
    # Get usage summary for last 7 days
    from datetime import datetime, timezone, timedelta
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

    result = await (
        supabase.table("usage_log")
        .select("client_id, action_type, tokens_used, duration_seconds")
        .gte("created_at", week_ago)
        .limit(1000)
        .execute()
    )

    if not result.data:
        await update.message.reply_text("No usage data in the last 7 days.")
        return

    # Aggregate by client
    from collections import defaultdict
    stats: dict[int | None, dict] = defaultdict(lambda: {"chats": 0, "builds": 0, "tokens": 0, "duration": 0})
    for r in result.data:
        cid = r.get("client_id")
        stats[cid]["tokens"] += r.get("tokens_used", 0)
        stats[cid]["duration"] += r.get("duration_seconds", 0)
        if r["action_type"] == "chat":
            stats[cid]["chats"] += 1
        elif r["action_type"] in ("build_queued", "build_completed"):
            stats[cid]["builds"] += 1

    # Resolve client names
    clients_result = await supabase.table("clients").select("id, name").execute()
    name_map = {c["id"]: c["name"] for c in (clients_result.data or [])}
    name_map[None] = "Admin"

    lines = ["Usage (last 7 days):\n"]
    for cid, s in sorted(stats.items(), key=lambda x: x[1]["tokens"], reverse=True):
        name = name_map.get(cid, f"client_{cid}")
        lines.append(
            f"{name}: {s['chats']} chats, {s['builds']} builds, "
            f"{s['tokens']:,} tokens, {s['duration']:.0f}s"
        )

    await update.message.reply_text("\n".join(lines))


# ── Chat (free text → Claude with repo context) ─────────────────

async def _load_repo_context_block(repo_name: str) -> str:
    """Load repo context and format as a prompt block."""
    block = ""
    try:
        supabase = await get_supabase()
        ctx = await context_loader.load_brainstorm_context(repo_name, supabase)

        if ctx["claude_md"]:
            block += f"\n--- PROJECT RULES ({repo_name}) ---\n{ctx['claude_md'][:1000]}\n"
        if ctx["status_md"]:
            block += f"\n--- CURRENT STATUS ({repo_name}) ---\n{ctx['status_md'][:800]}\n"
        if ctx["recent_commits"]:
            # Only last 5 commits
            commits = "\n".join(ctx["recent_commits"].splitlines()[:5])
            block += f"\n--- RECENT CHANGES ({repo_name}) ---\n{commits}\n"
        if ctx["file_tree"]:
            # Only top-level dirs + key files (cap at 1KB)
            block += f"\n--- FILES ({repo_name}) ---\n{ctx['file_tree'][:1000]}\n"
        if ctx["memory"]:
            block += f"\n--- MEMORY ({repo_name}) ---\n"
            for m in ctx["memory"][:5]:
                block += f"- {m['key']}: {m['value']}\n"
        if ctx["repo_config"].get("deploy_url"):
            block += f"\nLive at: {ctx['repo_config']['deploy_url']}\n"
        local_path = os.path.expanduser(ctx["repo_config"].get("local_path", ""))
        if local_path:
            block += f"\nLocal repo path: {local_path}\n"
        block += f"""
--- DEPLOY INSTRUCTIONS ---
When you make direct file edits (not via [ACTION:BUILD]):
1. cd to the repo's local path
2. git add the changed files, git commit with a clear message
3. git push origin main
4. Run: cd {local_path} && npx vercel --prod --yes 2>&1 | tail -5
5. Report the deploy URL back to the user
"""
    except Exception as e:
        logger.warning(f"Failed to load context for {repo_name}: {e}")
    return block


async def _build_system_prompt(user: dict, chat_id: str) -> str:
    """Build a context-aware system prompt — technical for admin, conversational for clients."""
    repo_name = get_active_repo(chat_id, user)

    if is_admin(user):
        # ── ADMIN PROMPT: senior CTO engineer ──
        base = f"""You are a senior CTO and software architect working with Musa at Wingmen. You think like a principal engineer — opinionated, decisive, and practical. Keep responses concise (Telegram).

## HOW YOU THINK

1. SIMPLEST SOLUTION FIRST: Always propose the simplest approach that solves the problem. No premature abstractions, no over-engineering. Three similar lines of code > a clever abstraction. A flat file > a database until you need queries.

2. EXISTING PATTERNS OVER NEW ONES: Before suggesting anything, consider what already exists in the codebase. Match existing naming, file structure, component patterns, and styling. Don't introduce new libraries, frameworks, or patterns unless there's a compelling reason.

3. PRODUCT THINKING: Don't just build what's asked — think about WHY it's being asked. Consider the end user, the business context, the demo timeline. Push back if something doesn't make sense. Suggest better alternatives.

4. ARCHITECTURAL OPINIONS: Have strong opinions, loosely held. When asked "should we do X or Y?", give a clear recommendation with reasoning, not a list of pros/cons. Example: "Go with X. It's simpler, matches your existing patterns, and you can always add Y later if needed."

5. SCOPE RUTHLESSLY: For demos, build the minimum that tells the story. For production, build the minimum that ships. Cut features aggressively. "What can we NOT build?" is more valuable than "what should we build?"

## CLARIFICATION — MANDATORY

When Musa gives notes, requirements, or vague requests:
- Do NOT immediately generate build specs
- Ask 2-3 sharp clarifying questions about intent, scope, and priority
- Then give YOUR recommendation with reasoning (don't just ask — lead)
- Example: "These notes mention WhatsApp. For the demo, I'd mock the UI with green bubbles — real WhatsApp API integration is a 2-week project and not worth it for Wednesday. Agree?"
- If Musa says "you decide", make the call and state it clearly

## TASK SIZING — MANDATORY

Large tasks MUST be broken into smaller, focused jobs:
- Split by page/component/feature — one job per logical unit
- Each job should touch 3-5 files maximum
- Jobs run sequentially within the same repo — later jobs see earlier changes
- The AI build agent works best with surgical, focused tasks
- Bad: "Make admin mobile-friendly" (one massive job)
- Good: 4 jobs — layout shell, then each page individually

## BUILDING

When intent is clear and Musa confirms:
- Include exactly ONE [ACTION:BUILD] block per message — never multiple
- After queuing it, say "Queued. Ready for the next one?" and wait
- When Musa says "next" or "go", include the next [ACTION:BUILD]
- This keeps responses fast and reliable
- Be specific: name exact files, components, functions
- Include acceptance criteria the agent can verify

[ACTION:BUILD] surgical description — exact files, exact changes, exact acceptance criteria [/ACTION]

For data changes: [ACTION:DATA] TABLE/OP/DATA/WHERE format [/ACTION]

NEVER include action blocks without Musa confirming first.

## COMMANDS
/screenshots, /preview, /undo, /digest, /mu, /jobs, /status

Projects: {', '.join(user['repos'])}
"""
    else:
        # ── CLIENT PROMPT: smart product advisor ──
        base = f"""You are {user['name']}'s dedicated project advisor from Wingmen. You're not a generic chatbot — you're a smart, experienced product person who happens to have an AI dev team behind you.

## YOUR PERSONALITY
- Warm, professional, proactive. Like a trusted business partner who happens to know tech.
- Never use jargon (no "API", "component", "deploy", "responsive"). Use plain language.
- Be opinionated — don't ask "what do you prefer?" when you know the right answer. Say "I'd recommend X because Y. Sound good?"
- Anticipate needs. If they ask to update one product, ask if there are others to update too.
- Celebrate wins. When something is done, be genuinely enthusiastic.

## HOW YOU THINK
- Their time is valuable. Minimize back-and-forth — ask the essential questions in one message, not spread across five.
- Most requests are simple (update text, change price, fix a typo). Handle these instantly with DATA actions.
- When something requires building, explain what will happen in their terms: "I'll get our dev team to add that. It usually takes about 10 minutes — I'll send you a screenshot when it's ready."
- If something doesn't make sense for their business, say so. "I can do that, but have you considered X instead? It might work better because..."

## HANDLING REQUESTS

1. LISTEN & CLARIFY: Ask 1-2 essential questions. Don't over-clarify obvious requests.
   - "Update nasi lemak price to $6" → just do it, don't ask "which size?"
   - "I want to redesign the menu" → clarify "What's not working with the current one?"

2. RECOMMEND: Give your opinion. Don't present options — present a recommendation.
   - Bad: "We could do A, B, or C. Which do you prefer?"
   - Good: "I'd go with A — it's the simplest and your customers are used to it. Want me to do that?"

3. CONFIRM: Brief summary + "Should I go ahead?"

4. EXECUTE: Only after they confirm. Choose the lightest action:
   - Price/text/toggle changes → [ACTION:DATA] (instant)
   - New setup/config → [ACTION:CONFIG]
   - New features/pages/fixes → [ACTION:BUILD] (takes ~10 min)

## ACTION FORMAT (user never sees these blocks)

[ACTION:DATA]
TABLE: products
OP: update
WHERE: id=42
DATA: {{"price": 6.00}}
[/ACTION]

[ACTION:CONFIG] description of setup needed [/ACTION]

[ACTION:BUILD] detailed technical spec for dev agent — specific files, components, acceptance criteria. Be surgical — 3-5 files max per action. [/ACTION]

## RULES
- NEVER include action blocks on the first message about a topic
- Text before action blocks = what user sees (friendly, non-technical)
- Text inside action blocks = for dev agent (technical, specific)
- Prefer DATA over BUILD — most things are data changes
- Questions and chat = just respond normally, no actions
- When builds complete, user gets a screenshot automatically

{user['name']}'s project{'s' if len(user['repos']) > 1 else ''}: {', '.join(user['repos'])}
"""

    if len(user["repos"]) > 1 and repo_name:
        base += f"Currently discussing: {repo_name}\n"

    if repo_name:
        base += await _load_repo_context_block(repo_name)

        # Add completed job history so bot doesn't duplicate work
        try:
            supabase = await get_supabase()
            jobs_result = await (
                supabase.table("jobs")
                .select("id, description, status")
                .eq("repo_name", repo_name)
                .in_("status", ["completed", "running", "queued"])
                .order("id", desc=True)
                .limit(10)
                .execute()
            )
            if jobs_result.data:
                base += "\n--- ALREADY BUILT/IN PROGRESS (do NOT rebuild these) ---\n"
                for j in jobs_result.data:
                    desc = j["description"].split("\n")[0][:100]
                    base += f"- [{j['status']}] Job #{j['id']}: {desc}\n"
        except Exception:
            pass

    return base


async def _parse_and_execute_actions(reply: str, user: dict, chat_id: str) -> str:
    """Parse action blocks from Claude's response and execute them.

    Supports 3 tiers:
    - [ACTION:DATA] — direct Supabase query (instant, e.g. update prices)
    - [ACTION:CONFIG] — provisioning (new storefront, DNS, etc.)
    - [ACTION:BUILD] — full code build pipeline
    """
    import re

    repo_name = get_active_repo(chat_id, user)

    # ── DATA actions (instant Supabase operations) ──
    data_match = re.search(r'\[ACTION:DATA\]\s*(.+?)\s*\[/ACTION\]', reply, re.DOTALL)
    if data_match:
        sql_or_op = data_match.group(1).strip()
        clean_reply = reply[:data_match.start()].strip()
        try:
            supabase = await get_supabase()
            # Parse structured operations (table, operation, data)
            result = await _execute_data_action(supabase, sql_or_op, repo_name)
            clean_reply += f"\n\n{result}"
            logger.info(f"Data action for {user['name']}: {sql_or_op[:100]}")
        except Exception as e:
            logger.error(f"Data action failed: {e}")
            clean_reply += "\n\nI tried to make that change but hit an issue. Let me flag this to the team."
        return clean_reply

    # ── CONFIG actions (provisioning) ──
    config_match = re.search(r'\[ACTION:CONFIG\]\s*(.+?)\s*\[/ACTION\]', reply, re.DOTALL)
    if config_match:
        config_op = config_match.group(1).strip()
        clean_reply = reply[:config_match.start()].strip()
        try:
            result = await _execute_config_action(config_op, user)
            clean_reply += f"\n\n{result}"
            logger.info(f"Config action for {user['name']}: {config_op[:100]}")
        except Exception as e:
            logger.error(f"Config action failed: {e}")
            clean_reply += "\n\nI'll need to get the team to handle this setup. I've flagged it!"
        return clean_reply

    # ── BUILD actions (code changes via Claude CLI) — supports multiple ──
    build_matches = list(re.finditer(r'\[ACTION:BUILD\]\s*(.+?)\s*\[/ACTION\]', reply, re.DOTALL))
    if build_matches:
        # Clean reply = everything before the first action block
        clean_reply = reply[:build_matches[0].start()].strip()

        if not repo_name:
            return clean_reply + "\n\nWhich project is this for? Just let me know!"

        queued_ids = []
        for match in build_matches:
            technical_desc = match.group(1).strip()
            if len(technical_desc) > 2000:
                technical_desc = technical_desc[:2000]

            try:
                config = context_loader.get_repo_config(repo_name)
                supabase = await get_supabase()
                result = await supabase.table("jobs").insert({
                    "repo_name": repo_name,
                    "description": technical_desc,
                    "status": "queued",
                    "priority": config["priority"],
                    "triggered_by": "telegram",
                    "client_id": user.get("client_id"),
                }).execute()

                if result.data:
                    job = result.data[0]
                    queued_ids.append(str(job["id"]))
                    logger.info(f"Build queued #{job['id']} for {user['name']}: {repo_name} — {technical_desc[:100]}")
            except Exception as e:
                logger.error(f"Failed to queue build: {e}")

        if queued_ids:
            ids = ", #".join(queued_ids)
            clean_reply += f"\n\nI've submitted {len(queued_ids)} task{'s' if len(queued_ids) > 1 else ''} (#{ids}). I'll update you on progress!"
        else:
            clean_reply += "\n\nI tried to submit those but hit an issue. Let me flag this to the team."

        return clean_reply

    return reply


async def _execute_data_action(supabase, operation: str, repo_name: str | None) -> str:
    """Execute a structured data operation on Supabase.

    Expected format (from Claude):
    TABLE: <table_name>
    OP: insert|update|delete
    WHERE: <column>=<value>  (for update/delete)
    DATA: {"key": "value", ...}
    """
    import json as json_mod

    lines = operation.strip().splitlines()
    parsed = {}
    for line in lines:
        if ":" in line:
            key, val = line.split(":", 1)
            parsed[key.strip().upper()] = val.strip()

    table = parsed.get("TABLE", "")
    op = parsed.get("OP", "").lower()
    data_str = parsed.get("DATA", "{}")
    where_str = parsed.get("WHERE", "")

    if not table or not op:
        return "I couldn't process that change — the format was unclear. Let me try again."

    try:
        data = json_mod.loads(data_str)
    except json_mod.JSONDecodeError:
        return "I had trouble understanding the data format. Could you describe what you want changed?"

    if op == "insert":
        result = await supabase.table(table).insert(data).execute()
        if result.data:
            return "Done! The change is live now."
        return "I tried but the update didn't go through. Let me check with the team."

    elif op == "update" and where_str:
        col, val = where_str.split("=", 1)
        result = await supabase.table(table).update(data).eq(col.strip(), val.strip()).execute()
        if result.data:
            return "Done! The change is live now."
        return "I couldn't find that record to update. Can you double-check the details?"

    elif op == "delete" and where_str:
        col, val = where_str.split("=", 1)
        result = await supabase.table(table).delete().eq(col.strip(), val.strip()).execute()
        return "Done! That's been removed."

    return "I wasn't sure how to process that. Let me flag it to the team."


async def _execute_config_action(operation: str, user: dict) -> str:
    """Execute a configuration/provisioning action.

    Handles: new storefront setup, domain configuration, etc.
    """
    # For now, queue as a manual task for Musa
    supabase = await get_supabase()
    await supabase.table("jobs").insert({
        "repo_name": "orchestrator",
        "description": f"CONFIG: {operation}",
        "status": "queued",
        "priority": 2,
        "triggered_by": "telegram",
        "client_id": user.get("client_id"),
    }).execute()

    return "I've flagged this for setup. You'll be notified once it's ready!"


_chat_semaphore = asyncio.Semaphore(10)  # Max 10 concurrent Claude calls


async def handle_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle free-text messages."""
    chat_id = str(update.effective_user.id)
    msg_id = update.message.message_id
    user_msg = update.message.text
    if not user_msg:
        return

    # Periodic cache cleanup
    _cleanup_caches()

    # Queue first — survives crashes
    await _queue_message(chat_id, msg_id, "text", user_msg)

    user = await resolve_user(chat_id)
    if not user:
        await handle_onboarding(update, chat_id)
        await _mark_processed(chat_id, msg_id)
        return

    await _process_message(update, user, chat_id, user_msg)
    await _mark_processed(chat_id, msg_id)


CLAUDE_BIN = os.path.expanduser("~/.local/bin/claude")
CLAUDE_ENV = {
    "HOME": os.path.expanduser("~"),
    "PATH": os.environ.get("PATH", "/usr/local/bin:/usr/bin:/bin"),
    "USER": os.environ.get("USER", ""),
    "SHELL": os.environ.get("SHELL", ""),
    "LANG": os.environ.get("LANG", ""),
}


async def _call_claude(prompt: str, *, tools: str = "", timeout: int = 300) -> str:
    """Call Claude CLI (one-shot) and return the text response. Returns empty string on failure."""
    args = [CLAUDE_BIN, "-p", prompt, "--output-format", "text"]
    if tools:
        args += ["--allowedTools", tools, "--permission-mode", "bypassPermissions"]

    proc = await asyncio.create_subprocess_exec(
        *args,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        env=CLAUDE_ENV,
    )
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.wait()
        logger.error(f"Claude CLI timed out after {timeout}s")
        return ""

    result = stdout.decode(errors="replace").strip()
    if not result:
        logger.error(f"Claude CLI empty output: {stderr.decode(errors='replace')[:200]}")
    return result


async def _call_claude_session(prompt: str, *, tools: str = "", timeout: int = 600) -> str:
    """Call Claude CLI with persistent session (admin only). Resumes existing session for full context."""
    global _admin_session_id, _admin_session_ts
    import uuid

    now = time.monotonic()

    # Check if session expired
    new_session = not _admin_session_id or (now - _admin_session_ts > ADMIN_SESSION_TTL)

    if new_session:
        _admin_session_id = str(uuid.uuid4())
        logger.info(f"Starting new admin Claude session: {_admin_session_id}")

    _admin_session_ts = now

    args = [CLAUDE_BIN, "-p", prompt, "--output-format", "text"]
    if new_session:
        # First call: create session with --session-id
        args += ["--session-id", _admin_session_id]
    else:
        # Subsequent calls: resume existing session
        args += ["--resume", _admin_session_id]
    if tools:
        args += ["--allowedTools", tools, "--permission-mode", "bypassPermissions"]

    proc = await asyncio.create_subprocess_exec(
        *args,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        env=CLAUDE_ENV,
    )
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.wait()
        logger.error(f"Claude session timed out after {timeout}s")
        # Reset session on timeout
        _admin_session_id = None
        return ""

    result = stdout.decode(errors="replace").strip()
    if not result:
        err = stderr.decode(errors="replace")[:200]
        logger.error(f"Claude session empty output: {err}")
        # If session is broken, reset for next call
        if "session" in err.lower() or "resume" in err.lower():
            _admin_session_id = None
    return result


# ── Agent dispatch functions ─────────────────────────────────────


async def _route_message(user_msg: str, user: dict, history: list[dict]) -> dict:
    """Use Router Agent to classify message intent. Falls back to 'chat' on failure."""
    role = "admin" if is_admin(user) else "client"
    prompt = build_router_prompt(user_msg, user["repos"], history, role=role)
    raw = await _call_claude(prompt, timeout=30)
    if not raw:
        return {"intent": "chat", "repo": None, "detail": user_msg}
    result = parse_router_response(raw)
    logger.info(f"Router: {user['name']} -> {result['intent']} (repo={result.get('repo')}, role={role})")
    return result


async def _run_brainstorm(update: Update, user: dict, chat_id: str, history: list[dict], user_msg: str) -> str:
    """Run the Brainstorm Agent and return the reply text."""
    repo_name = get_active_repo(chat_id, user)
    repo_context = ""
    if repo_name:
        repo_context = await _load_repo_context_block(repo_name)

    prompt = build_brainstorm_prompt(
        user=user,
        repo_context=repo_context,
        history=history,
        user_msg=user_msg,
    )

    # Admin gets persistent session (full context across messages)
    # Clients get one-shot (no session persistence)
    if is_admin(user):
        reply = await _call_claude_session(prompt, tools="Read,Write,Edit,Glob,Grep,Bash,WebFetch,WebSearch", timeout=600)
    else:
        reply = await _call_claude(prompt, tools="Read,Glob,Grep", timeout=300)

    if not reply:
        return "I had trouble processing that. Please try again."

    # Parse action blocks (DATA/CONFIG/BUILD)
    reply = await _parse_and_execute_actions(reply, user, chat_id)
    return reply


async def _run_audit(update: Update, user: dict, chat_id: str, repo_name: str, detail: str) -> str:
    """Run Auditor Agent, then handle results based on persona."""
    try:
        config = context_loader.get_repo_config(repo_name)
    except ValueError:
        return f"I don't have a repo called '{repo_name}' configured."

    deploy_url = config.get("deploy_url", "")
    repo_path = os.path.expanduser(config.get("local_path", ""))

    if not deploy_url:
        if is_admin(user):
            return f"No deploy URL configured for {repo_name}. Can't crawl pages without it."
        else:
            return "I can't check your site right now — the link isn't set up yet. I'll flag this for the team."

    # Get file tree for route list
    ctx_block = await _load_repo_context_block(repo_name)
    file_tree = ""
    if "--- FILES" in ctx_block:
        start = ctx_block.index("--- FILES")
        rest = ctx_block[start + 10:]
        end_offset = rest.find("\n---")
        if end_offset >= 0:
            file_tree = ctx_block[start:start + 10 + end_offset].strip()
        else:
            file_tree = ctx_block[start:].strip()

    # Persona-aware progress message
    if is_admin(user):
        await update.message.reply_text(f"Auditing {repo_name} — crawling all pages. This will take a few minutes...")
    else:
        await update.message.reply_text("Checking your site now, one moment...")

    prompt = build_auditor_prompt(
        deploy_url=deploy_url,
        repo_path=repo_path,
        file_tree=file_tree,
        detail=detail,
    )

    raw = await _call_claude(prompt, tools="WebFetch,WebSearch,Read,Glob,Grep,Bash", timeout=600)
    if not raw:
        if is_admin(user):
            return "Audit failed — Claude didn't return results. Try again?"
        else:
            return "I had trouble checking your site. Want me to try again?"

    issues, summary = parse_auditor_response(raw)
    high_conf = [i for i in issues if i.get("fix_confidence") == "high"]
    needs_decision = [i for i in issues if i.get("fix_confidence") != "high"]

    # ── ADMIN FLOW: auto-fix and raw report ──
    if is_admin(user):
        fix_results = []
        if high_conf:
            await update.message.reply_text(f"Found {len(issues)} issue(s). Auto-fixing {len(high_conf)} obvious one(s)...")
            for issue in high_conf:
                fix_reply = await _run_fix(repo_name, issue)
                fix_results.append(f"- {issue['description']}: {fix_reply}")
            if fix_results:
                push_result = await _call_claude(
                    f"Run these commands:\ncd {repo_path} && git push origin main 2>&1 | tail -3",
                    tools="Bash",
                    timeout=60,
                )
                logger.info(f"Batch push for {repo_name}: {push_result[:100]}")

        parts = [f"Audit complete for {repo_name} — {len(issues)} issue(s) found.\n"]
        if fix_results:
            parts.append(f"Fixed {len(fix_results)} automatically:")
            parts.extend(fix_results)
        if needs_decision:
            parts.append(f"\n{len(needs_decision)} need your call:")
            for i in needs_decision:
                sev = {"high": "!!!", "medium": "!!", "low": "!"}.get(i.get("severity", ""), "?")
                parts.append(f"{sev} {i['description']}")
                if i.get("suggested_fix"):
                    parts.append(f"   Suggestion: {i['suggested_fix']}")
        if not issues:
            parts = ["All pages look good — no issues found."]
        return "\n".join(parts)

    # ── CLIENT FLOW: translate and ask permission ──
    if high_conf:
        _pending_fixes[chat_id] = {"issues": high_conf, "repo": repo_name}

    translated = await _translate_audit_for_client(issues, repo_name)
    return translated


async def _run_fix(repo_name: str, issue: dict) -> str:
    """Run Fixer Agent for a single issue. Returns a one-line result."""
    try:
        config = context_loader.get_repo_config(repo_name)
    except ValueError:
        return "SKIP: repo not found"

    repo_path = os.path.expanduser(config.get("local_path", ""))

    prompt = build_fixer_prompt(issue=issue, repo_path=repo_path)
    result = await _call_claude(prompt, tools="Read,Edit,Write,Bash", timeout=120)

    if not result:
        return "SKIP: fixer returned no output"
    if result.strip().startswith("SKIP:"):
        return result.strip()
    return result.split("\n")[0][:200]  # First line, capped


async def _translate_audit_for_client(issues: list[dict], repo_name: str) -> str:
    """Translate a technical audit report into plain client-friendly language."""
    high_conf = [i for i in issues if i.get("fix_confidence") == "high"]
    needs_decision = [i for i in issues if i.get("fix_confidence") != "high"]

    if not high_conf and not needs_decision:
        return f"I checked your site ({repo_name}) and everything looks good!"

    prompt = f"""You are a friendly project advisor explaining website issues to a non-technical client.
Translate these findings into plain, warm language. NO file paths, NO code, NO severity ratings, NO technical jargon.

Group into two sections:
1. "I can fix these right now" — list the easy fixes in simple terms
2. "I'd like your input on" — list items that need their decision
"""

    if high_conf:
        prompt += f"\nEasy fixes ({len(high_conf)}):\n"
        for i in high_conf:
            prompt += f"- {i['description']}\n"
        prompt += '\nEnd with exactly: "Should I go ahead with the easy fixes?"'
    else:
        prompt += "\nNo easy fixes found.\n"

    if needs_decision:
        prompt += f"\nNeeds input ({len(needs_decision)}):\n"
        for i in needs_decision:
            prompt += f"- {i['description']}"
            if i.get("suggested_fix"):
                prompt += f" (suggestion: {i['suggested_fix']})"
            prompt += "\n"

    if not high_conf:
        prompt += '\nEnd with: "Let me know which of these you\'d like me to look into."'

    result = await _call_claude(prompt, timeout=60)
    return result or "I found some things on your site — let me put together a summary for you."


_CONFIRM_WORDS = {"yes", "yeah", "yep", "sure", "ok", "okay", "go ahead", "do it", "please", "proceed", "fix it", "go for it", "yea"}


async def _execute_pending_fixes(update: Update, user: dict, chat_id: str) -> str:
    """Execute pending audit fixes for a client and return a friendly summary."""
    pending = _pending_fixes.pop(chat_id, None)
    if not pending:
        return ""

    issues = pending["issues"]
    repo_name = pending["repo"]

    try:
        config = context_loader.get_repo_config(repo_name)
    except ValueError:
        return "I couldn't find the project to fix. Let me know if you'd like to try again."

    repo_path = os.path.expanduser(config.get("local_path", ""))

    await update.message.reply_text(f"On it! Fixing {len(issues)} thing{'s' if len(issues) != 1 else ''} now...")

    fixed_count = 0
    for issue in issues:
        result = await _run_fix(repo_name, issue)
        if not result.startswith("SKIP"):
            fixed_count += 1

    # Batch push
    if fixed_count > 0:
        await _call_claude(
            f"Run these commands:\ncd {repo_path} && git push origin main 2>&1 | tail -3",
            tools="Bash",
            timeout=60,
        )

    if fixed_count == len(issues):
        return f"All done! I've fixed all {fixed_count} issue{'s' if fixed_count != 1 else ''}. Take a look at your site and let me know if it looks better."
    elif fixed_count > 0:
        return f"I fixed {fixed_count} out of {len(issues)}. A couple were trickier than expected — I'll flag those for the team. Check your site and let me know!"
    else:
        return "I ran into some trouble with the fixes. Let me flag this for the team to look at."


async def _send_reply(update: Update, text: str, max_len: int = 4000):
    """Send a reply, splitting into multiple messages if needed."""
    if len(text) <= max_len:
        await update.message.reply_text(text)
        return

    # Split on double newlines (paragraph breaks) first, fall back to single newlines
    chunks = []
    current = ""
    for line in text.split("\n"):
        if len(current) + len(line) + 1 > max_len:
            if current:
                chunks.append(current.strip())
            current = line + "\n"
        else:
            current += line + "\n"
    if current.strip():
        chunks.append(current.strip())

    for i, chunk in enumerate(chunks):
        if i > 0:
            chunk = f"...({i+1}/{len(chunks)})\n\n{chunk}"
        await update.message.reply_text(chunk)


async def _process_message(update: Update, user: dict, chat_id: str, user_msg: str):
    """Core chat logic — routes to specialized agents via Router."""

    if len(user_msg) > MAX_MSG_LENGTH:
        await update.message.reply_text(f"Message too long (max {MAX_MSG_LENGTH} chars). Please shorten it.")
        return

    # Check usage limits
    limit_msg = await check_usage_limit(user, "chat")
    if limit_msg:
        await update.message.reply_text(limit_msg)
        return

    # Auto-detect repo from message or recent history
    if not get_active_repo(chat_id, user):
        for repo in user["repos"]:
            if repo.lower() in user_msg.lower():
                _active_repo[chat_id] = repo
                break
        if not get_active_repo(chat_id, user) and len(user["repos"]) == 1:
            _active_repo[chat_id] = user["repos"][0]
        if not get_active_repo(chat_id, user):
            inferred = await _load_active_repo(chat_id)
            if inferred and inferred in user["repos"]:
                _active_repo[chat_id] = inferred

    # If message contains a URL, auto-screenshot and enhance context
    url_enhanced = await handle_url_in_message(update, user, chat_id, user_msg)
    if url_enhanced:
        user_msg = url_enhanced

    # Load persisted history
    history = await get_history(chat_id)
    history.append({"role": "user", "content": user_msg})
    if len(history) > 20:
        history = history[-20:]

    # Persist user message
    await save_message(chat_id, "user", user_msg)

    # Check for pending fix confirmation (client flow)
    if chat_id in _pending_fixes and not is_admin(user):
        msg_lower = user_msg.lower().strip().rstrip("!.")
        if any(word in msg_lower for word in _CONFIRM_WORDS):
            async with _chat_semaphore:
                try:
                    await update.message.chat.send_action("typing")
                    reply = await _execute_pending_fixes(update, user, chat_id)
                    await save_message(chat_id, "assistant", reply)
                    await _send_reply(update, reply)
                except Exception as e:
                    logger.error(f"Pending fix error for {user['name']}: {e}")
                    try:
                        await update.message.reply_text("Something went wrong with the fixes. Let me try again later.")
                    except Exception:
                        pass
            return
        else:
            # Client said something other than confirmation — clear pending and route normally
            _pending_fixes.pop(chat_id, None)

    async with _chat_semaphore:
        start = time.monotonic()
        try:
            await update.message.chat.send_action("typing")

            # Step 1: Route the message
            route = await _route_message(user_msg, user, history)
            intent = route["intent"]

            # Override repo if router detected one
            if route.get("repo") and route["repo"] in user["repos"]:
                _active_repo[chat_id] = route["repo"]

            # Step 2: Dispatch to specialist agent with typing indicator
            async def _keep_typing():
                notified = False
                progress_msg = "Checking on that for you, one moment..." if not is_admin(user) else "Working on it — using tools to check things. Hang tight..."
                try:
                    elapsed = 0
                    while True:
                        await asyncio.sleep(8)
                        elapsed += 8
                        await update.message.chat.send_action("typing")
                        if not notified and elapsed >= 30:
                            notified = True
                            try:
                                await update.message.reply_text(progress_msg)
                            except Exception:
                                pass
                except asyncio.CancelledError:
                    pass
                except Exception:
                    pass

            typing_task = asyncio.create_task(_keep_typing())
            try:
                if intent == "audit":
                    repo = get_active_repo(chat_id, user) or route.get("repo")
                    if not repo:
                        reply = "Which project should I audit? Use /repo to set it."
                    else:
                        reply = await _run_audit(update, user, chat_id, repo, route.get("detail", user_msg))

                elif intent == "fix":
                    repo = get_active_repo(chat_id, user) or route.get("repo")
                    if not repo:
                        reply = "Which project? Use /repo to set it."
                    else:
                        # Direct fix: use fixer prompt but return full response (not truncated)
                        try:
                            config = context_loader.get_repo_config(repo)
                            repo_path = os.path.expanduser(config.get("local_path", ""))
                            deploy_url = config.get("deploy_url", "")
                        except ValueError:
                            repo_path = ""
                            deploy_url = ""
                        issue = {
                            "description": user_msg,  # full message including photo description + caption
                            "file_path": "",
                            "suggested_fix": route.get("detail", ""),
                        }
                        prompt = build_fixer_prompt(issue=issue, repo_path=repo_path)
                        if is_admin(user):
                            reply = await _call_claude_session(prompt, tools="Read,Write,Edit,Glob,Grep,Bash,WebFetch,WebSearch", timeout=300)
                        else:
                            reply = await _call_claude(prompt, tools="Read,Edit,Write,Bash", timeout=120)
                        if not reply:
                            reply = "I couldn't fix that. Could you give me more details?"
                        elif "SKIP" not in reply and deploy_url:
                            # Push changes
                            await _call_claude(
                                f"Run: cd {repo_path} && git push origin main 2>&1 | tail -3",
                                tools="Bash", timeout=60,
                            )
                            # Extract page route from fixer's PAGE: tag
                            page_route = "/"
                            for line in reply.strip().split("\n"):
                                line = line.strip()
                                if line.startswith("PAGE:"):
                                    page_route = line[5:].strip()
                                    break
                            # Clean the PAGE: tag from the user-facing reply
                            reply = "\n".join(l for l in reply.split("\n") if not l.strip().startswith("PAGE:"))
                            screenshot_url = deploy_url.rstrip("/") + page_route
                            screenshot_path = f"/tmp/fix_verify_{chat_id}.jpg"
                            await _call_claude(
                                f"Run: npx playwright screenshot {screenshot_url} {screenshot_path} --viewport-size 375,812 --wait-for-timeout 5000 2>&1 | tail -3",
                                tools="Bash", timeout=30,
                            )
                            if os.path.exists(screenshot_path):
                                try:
                                    with open(screenshot_path, "rb") as photo:
                                        await update.message.reply_photo(
                                            photo=photo,
                                            caption=f"Here's how it looks now after the fix.\n\nLive: {screenshot_url}"
                                        )
                                except Exception as e:
                                    logger.warning(f"Failed to send fix screenshot: {e}")
                                finally:
                                    os.unlink(screenshot_path)

                else:
                    # "chat", "build", "data" all go through brainstorm
                    reply = await _run_brainstorm(update, user, chat_id, history, user_msg)
            finally:
                typing_task.cancel()

            duration = time.monotonic() - start

            # Summarize long admin replies for Telegram
            if is_admin(user) and len(reply) > 4000:
                summary = await _call_claude(
                    f"Summarize this Claude Code output for a Telegram message. "
                    f"Keep the key information: what was done, what changed, any errors. "
                    f"Max 3000 chars. Use bullet points.\n\n{reply[:8000]}",
                    timeout=30,
                )
                telegram_reply = summary if summary else reply
            else:
                telegram_reply = reply

            # Persist full reply (not summarized)
            await save_message(chat_id, "assistant", reply)

            # Log usage
            repo = get_active_repo(chat_id, user) or ""
            await log_usage(user.get("client_id"), "chat", repo, 0, duration)

            # Send to Telegram (summarized for admin, full for clients)
            await _send_reply(update, telegram_reply)

        except asyncio.TimeoutError:
            logger.error(f"Chat timeout for {user['name']}")
            await update.message.reply_text("That took too long. Please try a shorter message.")

        except Exception as e:
            logger.error(f"Chat error for {user['name']}: {e}")
            try:
                await update.message.reply_text("Sorry, something went wrong. Please try again.")
            except Exception:
                logger.error(f"Failed to send error reply to {user['name']}")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo messages — save image and describe it in context for Claude."""
    chat_id = str(update.effective_user.id)
    msg_id = update.message.message_id

    photo = update.message.photo[-1] if update.message.photo else None
    if not photo:
        return

    caption = update.message.caption or ""
    await _queue_message(chat_id, msg_id, "photo", caption, photo.file_id)

    user = await resolve_user(chat_id)
    if not user:
        await handle_onboarding(update, chat_id)
        await _mark_processed(chat_id, msg_id)
        return

    caption = update.message.caption or ""

    tmp_path = None
    try:
        file = await context.bot.get_file(photo.file_id)
        import tempfile
        import shutil
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            tmp_path = tmp.name
            await file.download_to_drive(tmp_path)

        # Save persistent copy to ~/wingmen/photos/
        photos_dir = os.path.expanduser("~/wingmen/photos")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        user_name = user["name"].replace(" ", "_") if user else "unknown"
        persistent_path = os.path.join(photos_dir, f"{timestamp}_{user_name}_{msg_id}.jpg")
        shutil.copy2(tmp_path, persistent_path)
        logger.info(f"Photo saved to {persistent_path}")

        # Cleanup photos older than 3 months
        from datetime import timedelta
        cutoff = datetime.now() - timedelta(days=90)
        for old_file in Path(photos_dir).glob("*.jpg"):
            try:
                if datetime.fromtimestamp(old_file.stat().st_mtime) < cutoff:
                    old_file.unlink()
            except Exception:
                pass

        # Use Claude CLI with Read tool to describe the image
        describe_prompt = (
            f"Read the image file at {tmp_path} and describe it factually. "
            "Describe ONLY what you see — layout, elements, text, spacing, colors. "
            "Do NOT judge whether it looks good or bad. Do NOT say 'no issues' or 'looks fine'. "
            "Just describe the visual facts. Keep it under 200 words."
        )

        description = await _call_claude(describe_prompt, tools="Read", timeout=60)

        if not description:
            description = "(image received but couldn't be analyzed)"

        # If no caption, ask what they want fixed instead of guessing
        if not caption:
            await save_message(chat_id, "user", f"[User sent a photo: {description}]\n[Saved to: {persistent_path}]")
            reply = f"I can see: {description[:300]}\n\nWhat would you like me to do with this?"
            await save_message(chat_id, "assistant", reply)
            await update.message.reply_text(reply)
            await _mark_processed(chat_id, msg_id)
            return

        # Caption provided — process with context
        combined = f"[User sent a photo: {description}]\n[Saved to: {persistent_path}]\nUser's message: {caption}"
        await _process_message(update, user, chat_id, combined)
        await _mark_processed(chat_id, msg_id)

    except Exception as e:
        logger.error(f"Photo processing error: {e}")
        if caption:
            await _process_message(update, user, chat_id, f"[User sent a photo they want to discuss] {caption}")
        else:
            await update.message.reply_text("I received your photo. Could you describe what you'd like me to do with it?")
        await _mark_processed(chat_id, msg_id)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads — HTML, Excel, CSV, PDF."""
    chat_id = str(update.effective_user.id)
    msg_id = update.message.message_id
    user = await resolve_user(chat_id)

    doc = update.message.document
    if not doc:
        return

    caption = update.message.caption or ""
    file_name = doc.file_name or "file"
    file_ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    await _queue_message(chat_id, msg_id, "document", f"{file_name}: {caption}", doc.file_id)

    if not user:
        # For onboarding — save file context for provisioning
        await handle_onboarding(update, chat_id)
        await _mark_processed(chat_id, msg_id)
        return

    await update.message.chat.send_action("typing")

    tmp_path = None
    try:
        file = await context.bot.get_file(doc.file_id)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=f".{file_ext}", delete=False) as tmp:
            tmp_path = tmp.name
            await file.download_to_drive(tmp_path)

        content_summary = ""

        if file_ext in ("html", "htm"):
            # Read HTML directly
            with open(tmp_path, "r", errors="replace") as f:
                html = f.read()
            content_summary = f"[User sent an HTML file ({file_name}, {len(html)} chars)]\n\nHTML preview:\n{html[:3000]}"

        elif file_ext in ("xlsx", "xls"):
            # Parse Excel
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheets_data = []
            for sheet_name in wb.sheetnames[:3]:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=50, values_only=True):
                    rows.append(" | ".join(str(c) if c is not None else "" for c in row))
                sheets_data.append(f"Sheet: {sheet_name}\n" + "\n".join(rows[:50]))
            wb.close()
            content_summary = f"[User sent Excel file ({file_name})]\n\n" + "\n\n".join(sheets_data)

        elif file_ext == "csv":
            with open(tmp_path, "r", errors="replace") as f:
                csv_content = f.read()
            content_summary = f"[User sent CSV file ({file_name})]\n\n{csv_content[:3000]}"

        elif file_ext == "pdf":
            import pdfplumber
            with pdfplumber.open(tmp_path) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages[:5])
            content_summary = f"[User sent PDF ({file_name})]\n\n{text[:3000]}"

        elif file_ext in ("json",):
            with open(tmp_path, "r", errors="replace") as f:
                content = f.read()
            content_summary = f"[User sent JSON file ({file_name})]\n\n{content[:3000]}"

        else:
            # Try to read as text
            try:
                with open(tmp_path, "r", errors="replace") as f:
                    content = f.read()
                content_summary = f"[User sent file ({file_name})]\n\n{content[:2000]}"
            except Exception:
                content_summary = f"[User sent file ({file_name}) — binary format, cannot read directly]"

        if caption:
            content_summary += f"\n\nUser's message: {caption}"

        await _process_message(update, user, chat_id, content_summary)
        await _mark_processed(chat_id, msg_id)

    except Exception as e:
        logger.error(f"Document processing error: {e}")
        fallback = f"I received your file ({file_name})."
        if caption:
            fallback += f" {caption}"
        await _process_message(update, user, chat_id, fallback)
        await _mark_processed(chat_id, msg_id)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


async def handle_url_in_message(update: Update, user: dict, chat_id: str, user_msg: str) -> str | None:
    """Detect URLs in messages and auto-screenshot them for context.

    Returns enhanced message with screenshot description, or None if no URL found.
    """
    import re
    url_match = re.search(r'https?://[^\s<>\"]+', user_msg)
    if not url_match:
        return None

    url = url_match.group()

    try:
        await update.message.chat.send_action("typing")

        import tempfile
        ss_path = os.path.join(tempfile.gettempdir(), f"url_ref_{hash(url)}.png")
        proc = await asyncio.create_subprocess_exec(
            "npx", "playwright", "screenshot", url, ss_path,
            "--viewport-size", "375,812",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        await asyncio.wait_for(proc.communicate(), timeout=20)

        if proc.returncode == 0 and os.path.exists(ss_path):
            # Describe with Claude CLI
            claude_bin = os.path.expanduser("~/.local/bin/claude")
            safe_env = {k: v for k, v in os.environ.items() if k in {"PATH", "HOME", "USER", "SHELL", "LANG"}}
            safe_env["HOME"] = os.path.expanduser("~")
            safe_env["PATH"] = os.environ.get("PATH", "/usr/local/bin:/usr/bin:/bin")

            desc_proc = await asyncio.create_subprocess_exec(
                claude_bin, "-p",
                "Describe this website screenshot concisely — layout, colors, style, sections, typography. Keep it under 150 words.",
                "--files", ss_path,
                "--output-format", "text",
                "--tools", "",
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
                env=safe_env,
            )
            stdout, _ = await asyncio.wait_for(desc_proc.communicate(), timeout=30)
            description = stdout.decode(errors="replace").strip()

            os.unlink(ss_path)

            if description:
                enhanced = f"{user_msg}\n\n[Auto-captured screenshot of {url}: {description}]"
                return enhanced

    except Exception as e:
        logger.debug(f"URL screenshot failed for {url}: {e}")

    return None


async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle voice messages — download, transcribe via Claude, then process as text."""
    chat_id = str(update.effective_user.id)
    msg_id = update.message.message_id
    user = await resolve_user(chat_id)

    voice = update.message.voice or update.message.audio
    if not voice:
        return

    await _queue_message(chat_id, msg_id, "voice", "", voice.file_id)

    if not user:
        await handle_onboarding(update, chat_id)
        await _mark_processed(chat_id, msg_id)
        return

    # Step 1: Download and transcribe
    tmp_path = None
    try:
        file = await context.bot.get_file(voice.file_id)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            tmp_path = tmp.name
            await file.download_to_drive(tmp_path)

        transcription = await asyncio.to_thread(_whisper_transcribe, tmp_path)
    except Exception as e:
        logger.error(f"Voice transcription error: {e}")
        await update.message.reply_text("I had trouble with the audio. Could you try again or type it instead?")
        return
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not transcription:
        await update.message.reply_text("I couldn't make out what you said. Could you try again?")
        return

    # Step 2: Show transcription and process as chat
    await update.message.reply_text(f"\U0001f3a4 I heard: \"{transcription}\"")
    await _process_message(update, user, chat_id, transcription)
    await _mark_processed(chat_id, msg_id)


async def handle_unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_user.id)
    user = await resolve_user(chat_id)
    if not user:
        await update.message.reply_text("Not registered. Contact the admin.")
        return
    await update.message.reply_text("Unknown command. Send /start for help.")


# ── Main ─────────────────────────────────────────────────────────

def main():
    token = os.environ["TELEGRAM_BOT_TOKEN"]

    request = HTTPXRequest(
        read_timeout=30,
        write_timeout=30,
        connect_timeout=15,
        pool_timeout=10,
        connection_pool_size=20,
    )
    get_updates_request = HTTPXRequest(
        read_timeout=30,
        write_timeout=30,
        connect_timeout=15,
        pool_timeout=10,
    )
    app = (
        Application.builder()
        .token(token)
        .request(request)
        .get_updates_request(get_updates_request)
        .build()
    )

    # Universal commands
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_start))
    app.add_handler(CommandHandler("id", cmd_id))
    app.add_handler(CommandHandler("repo", cmd_repo))
    app.add_handler(CommandHandler("build", cmd_build))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("jobs", cmd_jobs))
    app.add_handler(CommandHandler("repos", cmd_repos))

    # Admin-only commands
    app.add_handler(CommandHandler("pause", cmd_pause))
    app.add_handler(CommandHandler("cancel", cmd_cancel))
    app.add_handler(CommandHandler("priority", cmd_priority))
    app.add_handler(CommandHandler("addclient", cmd_addclient))
    app.add_handler(CommandHandler("linkrepo", cmd_linkrepo))
    app.add_handler(CommandHandler("clients", cmd_clients))
    app.add_handler(CommandHandler("usage", cmd_usage))
    app.add_handler(CommandHandler("upgrade", cmd_upgrade))
    app.add_handler(CommandHandler("mu", cmd_plan_usage))  # admin: plan usage
    app.add_handler(CommandHandler("undo", cmd_undo))
    app.add_handler(CommandHandler("preview", cmd_preview))
    app.add_handler(CommandHandler("digest", cmd_digest))
    app.add_handler(CommandHandler("screenshots", cmd_screenshots))

    # Free text + voice → brainstorm
    app.add_handler(MessageHandler(filters.COMMAND, handle_unknown))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.VOICE | filters.AUDIO, handle_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_chat))

    # Recover messages, pre-load Whisper, and start nervous system
    async def post_init(application):
        await _recover_unprocessed(application)
        try:
            await asyncio.to_thread(_get_whisper)
            logger.info("Whisper model pre-loaded")
        except Exception as e:
            logger.warning(f"Failed to pre-load Whisper model: {e}")

        # Start nervous system scheduler
        try:
            from apscheduler.schedulers.asyncio import AsyncIOScheduler
            from nervous_system.brain_sync import brain_sync
            from nervous_system.morning_brief import morning_brief
            from nervous_system.weekly_digest import weekly_digest
            from nervous_system.memory_sync import memory_sync
            from nervous_system.session_compress import session_compress

            supabase_client = await get_supabase()
            admin_chat_id = os.environ.get("MUSA_TELEGRAM_ID", "286619815")

            scheduler = AsyncIOScheduler(timezone="Asia/Singapore")

            scheduler.add_job(
                brain_sync, "interval", hours=4, args=[supabase_client, application.bot, admin_chat_id],
                id="brain_sync", replace_existing=True,
            )
            scheduler.add_job(
                morning_brief, "cron", hour=6, args=[supabase_client, application.bot, admin_chat_id],
                id="morning_brief", replace_existing=True,
            )
            scheduler.add_job(
                weekly_digest, "cron", day_of_week="sun", hour=21, args=[supabase_client, application.bot, admin_chat_id],
                id="weekly_digest", replace_existing=True,
            )
            scheduler.add_job(
                memory_sync, "cron", hour=0, args=[supabase_client],
                id="memory_sync", replace_existing=True,
            )
            scheduler.add_job(
                session_compress, "cron", hour=2, args=[supabase_client],
                id="session_compress", replace_existing=True,
            )

            scheduler.start()
            logger.info("Nervous system scheduler started — 5 tasks registered")

            # Run brain_sync immediately on startup
            asyncio.create_task(brain_sync(supabase_client, application.bot, admin_chat_id))
        except Exception as e:
            logger.error(f"Failed to start nervous system scheduler: {e}")

    app.post_init = post_init

    async def error_handler(update, context):
        import httpx
        err = context.error
        # Suppress noisy network errors — these auto-recover via polling retry
        if isinstance(err, (httpx.ConnectError, httpx.ReadError, httpx.ReadTimeout, httpx.ConnectTimeout)):
            logger.warning(f"Network blip: {type(err).__name__}")
            return
        logger.error(f"Unhandled error: {err}")

    app.add_error_handler(error_handler)

    logger.info("Wingmen CTO Bot starting (long-polling mode)...")
    app.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=False,
        poll_interval=1.0,
        timeout=30,
    )


if __name__ == "__main__":
    main()
